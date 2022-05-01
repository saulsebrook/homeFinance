import openpyxl

budget = openpyxl.load_workbook('/Users/saulsebrook/Documents/Aulsebrook Budget.xlsx', data_only=True)

    
def get_allocations():  
    allocations = {}
    sheet = budget['Budget2022']
    x = list(sheet.columns)[9]
    y = list(sheet.columns)[8]
    for i in range(16, 31):
        allocations[y[i].value] = x[i].value
    
    return allocations

def get_income():
    income = {}
    sheet = budget['Budget2022']
    x = list(sheet.columns)[9]  
    y = list(sheet.columns)[8]
    for i in range(7, 9):
        income[y[i].value] = x[i].value
    
    return income

def get_presents_allocation():
    presents = {}
    sheet = budget['Budget2022']
    x = list(sheet.columns)[14]
    y = list(sheet.columns)[13]
    for i in range(7, 27):
        presents[y[i].value] = x[i].value
    
    return presents

def amend_wage(person, wage):
    sheet = budget['Budget2022']
    for rownum in range(8, 10):
        income = sheet.cell(row=rownum, column=9).value 
        if income == person:
            sheet.cell(row=rownum, column=10).value = wage

    budget.save('/Users/saulsebrook/Documents/Aulsebrook Budget.xlsx')

i = get_income()
print(list(i.keys())[0])


""" 

sheet = budget['Budget2022']
for rownum in range(8, 10):
  income = sheet.cell(row=rownum, column=9).value 
  if income in i:
    sheet.cell(row=rownum, column=10).value = 2350


budget.save('/Users/saulsebrook/Documents/Aulsebrook Budget.xlsx') """